import json
import os
import openpyxl
from PIL import Image, ImageDraw

excel_file = '../mapping/MAPPING DELTA.xlsx'
out_json = 'data/map_drafts/auto_mapped_delta_floors_draft.json'
out_report = 'data/map_drafts/auto_mapping_report.md'
previews_dir = 'data/map_drafts/previews'

os.makedirs(previews_dir, exist_ok=True)

floor_config = {
    'Tầng 1': {'num': 1, 'img_path': '../mapping/floor1.png'},
    'Tấng 2': {'num': 2, 'img_path': None},
    'Tầng 3': {'num': 3, 'img_path': '../mapping/floor3.png'},
    'Tầng 4': {'num': 4, 'img_path': None},
}

wb = openpyxl.load_workbook(excel_file, data_only=True)

mapped_floors = []
report_lines = ["# Delta Auto-Mapping Report\n"]

report_lines.append("**WARNING: This is an auto-generated draft and must be visually reviewed before replacing production map data.**\n")

total_auto_mapped = 0
total_skipped = 0
suspicious_items = []

for sheet_name in wb.sheetnames:
    if sheet_name not in floor_config:
        continue
        
    ws = wb[sheet_name]
    f_info = floor_config[sheet_name]
    f_num = f_info['num']
    img_path = f_info['img_path']
    
    report_lines.append(f"## Floor {f_num} ({sheet_name})")
    
    if not img_path or not os.path.exists(img_path):
        report_lines.append("- **Status:** `missing_floorplan_image`")
        
        # Count rooms roughly
        count = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip():
                    count += 1
        
        report_lines.append(f"- **Rooms detected in Excel:** {count}")
        report_lines.append("- **Auto-mapped:** 0 (Skipped due to missing image)\n")
        total_skipped += count
        
        mapped_floors.append({
            "floor": f_num,
            "image": None,
            "rooms": []
        })
        continue

    # Load image to get dimensions
    img = Image.open(img_path).convert('RGBA')
    img_w, img_h = img.size
    
    report_lines.append(f"- **Image Dimensions:** {img_w}x{img_h}")
    report_lines.append(f"- **Excel Grid:** {ws.max_row} rows x {ws.max_column} columns")
    
    # Calculate non-uniform grid sizes
    # Default widths/heights if not specified
    default_w = 10
    default_h = 15
    
    col_widths = []
    for c in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        cd = ws.column_dimensions[col_letter]
        # Custom width is usually in characters, roughly multiply by 7 for "pixels"
        w = (cd.width * 7) if cd.width else default_w
        col_widths.append(w)
        
    row_heights = []
    for r in range(1, ws.max_row + 1):
        rd = ws.row_dimensions[r]
        h = rd.height if rd.height else default_h
        row_heights.append(h)
        
    total_w = sum(col_widths)
    total_h = sum(row_heights)
    
    if total_w == 0 or total_h == 0:
        report_lines.append("- **Error:** Invalid excel grid sizes.\n")
        continue
        
    scale_x = img_w / total_w
    scale_y = img_h / total_h
    
    # Compute cumulative lines
    x_lines = [0]
    for w in col_widths:
        x_lines.append(x_lines[-1] + w)
        
    y_lines = [0]
    for h in row_heights:
        y_lines.append(y_lines[-1] + h)
        
    rooms = []
    processed_merged_ranges = set()
    floor_mapped_count = 0
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            
            if val and str(val).strip():
                code = str(val).strip()
                
                # Determine min/max row/col
                min_r, max_r, min_c, max_c = row, row, col, col
                is_merged = False
                
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        if merged_range.coord in processed_merged_ranges:
                            # Already processed this merged cell
                            code = None 
                            break
                        processed_merged_ranges.add(merged_range.coord)
                        min_r, min_c, max_r, max_c = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                        is_merged = True
                        break
                        
                if code is None:
                    continue
                    
                # Calculate pixel bbox
                px_min_x = x_lines[min_c - 1] * scale_x
                px_max_x = x_lines[max_c] * scale_x
                px_min_y = y_lines[min_r - 1] * scale_y
                px_max_y = y_lines[max_r] * scale_y
                
                # Sanity clip
                px_min_x = max(0, min(px_min_x, img_w))
                px_max_x = max(0, min(px_max_x, img_w))
                px_min_y = max(0, min(px_min_y, img_h))
                px_max_y = max(0, min(px_max_y, img_h))
                
                width = px_max_x - px_min_x
                height = px_max_y - px_min_y
                
                cx = (px_min_x + px_max_x) / 2
                cy = (px_min_y + px_max_y) / 2
                
                conf = "medium"
                notes = []
                
                if width < 15 or height < 15:
                    conf = "needs_review"
                    notes.append("Extremely small bounding box")
                    
                if width > img_w * 0.4 or height > img_h * 0.4:
                    conf = "needs_review"
                    notes.append("Unusually large bounding box")
                    
                if not is_merged and len(code) > 4:
                    conf = "low"
                    notes.append("Single cell for potentially large room")
                    
                rooms.append({
                    "room_code": code,
                    "item_type": "room",
                    "bbox": {
                        "min_x": int(px_min_x),
                        "min_y": int(px_min_y),
                        "max_x": int(px_max_x),
                        "max_y": int(px_max_y)
                    },
                    "center_x": int(cx),
                    "center_y": int(cy),
                    "label_x": int(cx),
                    "label_y": int(cy) - 10,
                    "source": "auto_mapped_from_xlsx_and_floorplan",
                    "geometry_confidence": conf,
                    "review_notes": "; ".join(notes)
                })
                
                if conf == "needs_review":
                    suspicious_items.append(f"F{f_num} - {code}: {'; '.join(notes)}")
                    
                floor_mapped_count += 1
                total_auto_mapped += 1

    mapped_floors.append({
        "floor": f_num,
        "image": os.path.basename(img_path),
        "image_width": img_w,
        "image_height": img_h,
        "rooms": rooms
    })
    
    report_lines.append(f"- **Auto-mapped rooms:** {floor_mapped_count}")
    
    # Generate Preview
    draw = ImageDraw.Draw(img)
    for r in rooms:
        b = r['bbox']
        color = "green"
        if r['geometry_confidence'] == "low":
            color = "orange"
        elif r['geometry_confidence'] == "needs_review":
            color = "red"
            
        # Draw semi-transparent fill
        draw.rectangle([b['min_x'], b['min_y'], b['max_x'], b['max_y']], outline=color, width=3)
        draw.text((r['label_x'], r['label_y']), r['room_code'], fill="blue")
        
    preview_file = os.path.join(previews_dir, f'floor{f_num}_auto_overlay.png')
    img.save(preview_file)
    report_lines.append(f"- **Preview:** `{preview_file}`\n")

report_lines.append("## Summary")
report_lines.append(f"- **Total Auto-mapped:** {total_auto_mapped}")
report_lines.append(f"- **Total Skipped (no image):** {total_skipped}")

report_lines.append("\n## Suspicious Items (needs_review)")
for s in suspicious_items:
    report_lines.append(f"- {s}")
if not suspicious_items:
    report_lines.append("- None")

with open(out_report, 'w', encoding='utf-8') as f:
    f.write("\n".join(report_lines))
    
with open(out_json, 'w', encoding='utf-8') as f:
    json.dump(mapped_floors, f, ensure_ascii=False, indent=2)

print("Auto-mapping complete.")
print(f"Total mapped: {total_auto_mapped}")
print(f"Suspicious: {len(suspicious_items)}")
