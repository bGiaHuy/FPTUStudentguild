import json
import os
import openpyxl
import re
import unicodedata

excel_file = '../mapping/MAPPING DELTA.xlsx'
out_json = 'data/map_drafts/delta_manual_measurements_draft.json'

def normalize_string(s):
    if not s:
        return "unknown"
    s = str(s)
    # Remove accents
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return s.strip('-')

def determine_type(display_name):
    norm = normalize_string(display_name)
    if 'cau-thang' in norm or 'stair' in norm: return 'stair'
    if 've-sinh' in norm or 'wc' in norm: return 'restroom'
    if 'hanh-lang' in norm or 'hallway' in norm: return 'hallway'
    if 'gieng-troi' in norm or 'skylight' in norm: return 'skylight'
    if 'sanh' in norm or 'lobby' in norm: return 'lobby'
    if 'cua-ra' in norm or 'entrance' in norm: return 'entrance'
    if norm == '---': return 'room' # '???' becomes '---' after normalize_string
    if 'tuong' in norm or 'wall' in norm: return 'wall'
    if norm.isnumeric() or 'phong' in norm or len(norm) <= 5: return 'room'
    return 'other'

if not os.path.exists(excel_file):
    print(f"File not found: {excel_file}")
    exit(1)

wb = openpyxl.load_workbook(excel_file, data_only=True)

floor_config = {
    'Tầng 1': 1,
    'Tầng 3': 3,
}

items = []
seen_ids = set()

for sheet_name, f_num in floor_config.items():
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    processed_merged = set()
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            
            if val and str(val).strip():
                display_name = str(val).strip()
                
                # Check merged cells
                is_merged = False
                min_r, min_c = row, col
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        if merged_range.coord in processed_merged:
                            display_name = None
                            break
                        processed_merged.add(merged_range.coord)
                        min_r, min_c = merged_range.min_row, merged_range.min_col
                        is_merged = True
                        break
                        
                if not display_name:
                    continue
                    
                item_type = determine_type(display_name)
                norm_name = normalize_string(display_name)
                
                # generate ID
                base_id = f"DELTA-F{f_num}-{item_type}-{norm_name}-r{min_r:02d}-c{min_c:02d}"
                
                count = 1
                item_id = f"{base_id}-{count:02d}"
                while item_id in seen_ids:
                    count += 1
                    item_id = f"{base_id}-{count:02d}"
                seen_ids.add(item_id)
                
                items.append({
                    "item_id": item_id,
                    "display_name": display_name,
                    "room_code": display_name,
                    "floor": f_num,
                    "item_type": item_type,
                    "source_row": min_r,
                    "source_col": min_c,
                    "source_sheet": sheet_name,
                    "bbox": {
                        "min_x": 0,
                        "min_y": 0,
                        "max_x": 0,
                        "max_y": 0
                    },
                    "center_x": 0,
                    "center_y": 0,
                    "label_x": 0,
                    "label_y": 0,
                    "click_radius": 20,
                    "annotation_status": "unassigned",
                    "source": "manual_measurement_tool",
                    "confidence": "needs_review"
                })

os.makedirs('data/map_drafts', exist_ok=True)
if not os.path.exists(out_json):
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"Generated new draft JSON at {out_json}")
else:
    print(f"File {out_json} already exists. Not overwriting to preserve IDs.")
    
# Debug: check duplicates
name_counts = {}
for i in items:
    name_counts[i['display_name']] = name_counts.get(i['display_name'], 0) + 1

dup_count = sum(1 for v in name_counts.values() if v > 1)
print(f"Total items: {len(items)}")
print(f"Total duplicate display_names: {dup_count}")
