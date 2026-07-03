import os
import re
import json
from collections import defaultdict
from openpyxl import load_workbook

FILE_PATH = os.path.join("..", "mapping", "MAPPING DELTA (1).xlsx")
OUTPUT_JSON = os.path.join("data", "delta_floors.json")
REPORT_JSON = os.path.join("data", "delta_xlsx_conversion_report.json")

CELL_WIDTH = 40
CELL_HEIGHT = 24

def normalize_floor_name(name):
    name = name.strip()
    if "Tấng" in name:
        name = name.replace("Tấng", "Tầng")
    return name

def classify_label(label):
    label_str = str(label).strip()
    label_lower = label_str.lower()
    
    if label_lower in ["tầng 1", "tầng 2", "tầng 3", "tầng 4", "tầng 5"]:
        return "header"
    if label_lower == "icpdp":
        return "office"
    if label_lower == "startup":
        return "research_center"
    if label_lower.startswith("cửa "):
        return "door"
    if re.match(r"^[A-Za-z]\d{3}$", label_str):
        return "classroom_lab"
    if re.match(r"^\d{3}[A-Za-z]?$", label_str):
        return "room"
    if label_lower in ["nvs nam", "nvs nữ", "wc", "nhà vệ sinh"]:
        return "toilet"
    if "cầu thang" in label_lower:
        return "stair"
    if label_lower in ["cửa", "cửa ra", "cửa ra vào"]:
        return "door"
    if "thư viện" in label_lower:
        return "library"
    if "kỹ thuật" in label_lower:
        return "technical_room"
    if label_lower in ["sảnh", "sanh"]:
        return "lobby"
    if label_lower in ["đường đi/ trống", "đường đi", "trống"]:
        return "hallway_or_empty"
    if "???" in label_str:
        return "room"
    if label_lower == "tường":
        return "wall"
    return "needs_review"

def get_codes_and_metadata(val, item_type):
    val_str = str(val).strip()
    display_code = val_str
    aliases = [val_str]
    
    if val_str.lower() == "icpdp":
        display_code = "Phòng Hợp tác Quốc tế & Phát triển Cá nhân (ICPDP)"
        aliases = ['icpdp', 'ICPDP', 'Phòng Hợp tác Quốc tế', 'Hợp tác Quốc tế & Phát triển Cá nhân']
        return val_str, display_code, aliases
        
    if val_str.lower() == "startup":
        display_code = "Startup / Research Center"
        aliases = ['startup', 'Startup', 'Research Center', 'Startup Research Center']
        return val_str, display_code, aliases
        
    if item_type in ["room", "classroom_lab"]:
        if re.match(r"^\d{3}[A-Za-z]?$", val_str):
            display_code = f"DE-{val_str}"
            aliases = [val_str, display_code]
            return val_str, display_code, aliases
            
    return val_str, val_str, [val_str]

def process_excel():
    print(f"Loading {FILE_PATH}...")
    wb = load_workbook(FILE_PATH, data_only=True)
    
    building_data = {
        "building": {
            "name": "MAPPING DELTA",
            "code": "DELTA"
        },
        "floors": []
    }
    
    report_data = {
        "summary": {},
        "ignored_labels": [],
        "unknown_labels": [],
        "needs_classification_review": [],
        "duplicate_room_codes_global": [],
        "duplicate_room_codes_same_floor": {}
    }
    
    global_labels = defaultdict(list)
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'DEMO':
            continue
            
        normalized_floor = normalize_floor_name(sheet_name)
        floor_num_match = re.search(r'\d+', normalized_floor)
        floor_num = int(floor_num_match.group()) if floor_num_match else 0
        
        ws = wb[sheet_name]
        floor_items = []
        floor_counts = {"total": 0, "needs_review": 0, "unknown": 0, "ignored": 0}
        floor_labels_seen = defaultdict(list)
        
        merged_ranges_map = {}
        for mr in ws.merged_cells.ranges:
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    merged_ranges_map[(row, col)] = mr
                    
        visited_merged = set()
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    val = str(cell.value).strip()
                    coord = cell.coordinate
                    
                    mr = merged_ranges_map.get((cell.row, cell.column))
                    if mr:
                        range_str = str(mr)
                        if range_str in visited_merged:
                            continue
                        visited_merged.add(range_str)
                        min_col, min_row, max_col, max_row = mr.min_col, mr.min_row, mr.max_col, mr.max_row
                        source_range = range_str
                    else:
                        min_col, min_row, max_col, max_row = cell.column, cell.row, cell.column, cell.row
                        source_range = coord
                        
                    # Manual Decision Corrections
                    # Rename C7:D7 219 to 219A
                    if normalized_floor == "Tầng 2" and val == "219" and source_range == "C7:D7":
                        val = "219A"
                        
                    x1 = (min_col - 1) * CELL_WIDTH
                    y1 = (min_row - 1) * CELL_HEIGHT
                    x2 = max_col * CELL_WIDTH
                    y2 = max_row * CELL_HEIGHT
                    
                    center_x = (x1 + x2) / 2
                    center_y = (y1 + y2) / 2
                    
                    item_type = classify_label(val)
                    
                    if item_type == "header":
                        floor_counts["ignored"] += 1
                        report_data["ignored_labels"].append({"label": val, "floor": normalized_floor, "cell": coord})
                        continue
                    
                    room_code, display_code, aliases = get_codes_and_metadata(val, item_type)
                    
                    safe_label = re.sub(r'[^A-Za-z0-9]', '_', val)
                    item_id = f"DELTA-F{floor_num}-{safe_label}-{source_range.replace(':', '_')}"
                    
                    searchable = item_type in ["room", "classroom_lab", "library", "facility", "office", "technical_room", "research_center"]
                    highlightable = item_type in ["room", "classroom_lab", "library", "facility", "office", "technical_room", "research_center", "toilet", "stair", "door", "lobby"]
                    
                    item_data = {
                        "item_id": item_id,
                        "room_code": room_code,
                        "display_code": display_code,
                        "aliases": aliases,
                        "label": val,
                        "type": item_type,
                        "center_x": center_x,
                        "center_y": center_y,
                        "bbox": {
                            "x1": x1,
                            "y1": y1,
                            "x2": x2,
                            "y2": y2
                        },
                        "source_sheet": sheet_name,
                        "normalized_floor_name": normalized_floor,
                        "source_cell": coord,
                        "source_range": source_range,
                        "coordinate_source": "excel_approx",
                        "needs_manual_geometry_review": False,
                        "searchable": searchable,
                        "highlightable": highlightable
                    }
                    
                    if item_type == "unknown":
                        item_data["needs_human_confirmation"] = True
                        
                    # Manual Decision: 440 on Tầng 2
                    if normalized_floor == "Tầng 2" and val == "440":
                        item_data["needs_human_confirmation"] = True
                        item_data["note"] = "Needs later verification because 440 also appears on Tầng 4"
                        
                    if item_type == "door" and val.lower().startswith("cửa "):
                        parts = val.split(maxsplit=1)
                        if len(parts) > 1:
                            item_data["linked_room_code"] = parts[1]
                            
                    floor_counts["total"] += 1
                    
                    if item_type == "needs_review":
                        floor_counts["needs_review"] += 1
                        report_data["needs_classification_review"].append(item_data)
                    elif item_type == "unknown":
                        floor_counts["unknown"] += 1
                        report_data["unknown_labels"].append(item_data)
                        
                    floor_labels_seen[val].append(coord)
                    global_labels[val].append(normalized_floor)
                        
                    floor_items.append(item_data)
        
        # Check floor duplicates
        floor_dupes = {k: v for k, v in floor_labels_seen.items() if len(v) > 1 and classify_label(k) in ["room", "classroom_lab"]}
        if floor_dupes:
            report_data["duplicate_room_codes_same_floor"][normalized_floor] = floor_dupes
        
        building_data["floors"].append({
            "floor_number": floor_num,
            "floor_name": normalized_floor,
            "items": floor_items,
            "grid_metadata": {
                "cell_width": CELL_WIDTH,
                "cell_height": CELL_HEIGHT
            }
        })
        
        report_data["summary"][normalized_floor] = floor_counts

    # Check global duplicates
    global_dupes = {k: list(set(v)) for k, v in global_labels.items() if len(v) > 1 and classify_label(k) in ["room", "classroom_lab"]}
    report_data["duplicate_room_codes_global"] = global_dupes

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(building_data, f, ensure_ascii=False, indent=2)
        
    with open(REPORT_JSON, "w", encoding="utf-8") as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)

    print(f"Conversion complete. Data saved to {OUTPUT_JSON} and {REPORT_JSON}")

if __name__ == '__main__':
    process_excel()
