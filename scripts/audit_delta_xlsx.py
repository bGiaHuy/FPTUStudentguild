import os
import re
from openpyxl import load_workbook

FILE_PATH = os.path.join("..", "mapping", "MAPPING DELTA (1).xlsx")

def normalize_floor_name(name):
    name = name.strip()
    if "Tấng" in name:
        name = name.replace("Tấng", "Tầng")
    return name

def classify_label(label):
    if not isinstance(label, str):
        label = str(label)
    label = label.strip()
    label_lower = label.lower()
    
    # C201, C302, etc. -> classroom/lab
    if re.match(r"^[A-Za-z]\d{3}$", label):
        return "classroom/lab"
    # numeric room codes like 220, 321 -> room
    if re.match(r"^\d{3}$", label):
        return "room"
    # Toilet
    if label_lower in ["nvs nam", "nvs nữ", "wc", "nhà vệ sinh"]:
        return "toilet"
    # Stair
    if "cầu thang" in label_lower:
        return "stair"
    # Door
    if label_lower in ["cửa", "cửa ra", "cửa ra vào"]:
        return "door"
    # Library
    if "thư viện" in label_lower:
        return "library"
    # Technical
    if "kỹ thuật" in label_lower:
        return "technical_room"
    if "???" in label:
        return "room"
    if label_lower == "tường":
        return "wall"
    
    return "needs_review"

def audit_xlsx():
    print(f"Loading {FILE_PATH}...")
    try:
        wb = load_workbook(FILE_PATH, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    print("\n1. Workbook sheet names:")
    print(wb.sheetnames)
    
    print("\n2. Normalized floor names:")
    for name in wb.sheetnames:
        print(f" - {name} -> {normalize_floor_name(name)}")
        
    for sheet_name in wb.sheetnames:
        if sheet_name == 'DEMO':
            continue
            
        print(f"\n{'='*50}")
        print(f"AUDITING SHEET: {sheet_name} (Normalized: {normalize_floor_name(sheet_name)})")
        print(f"{'='*50}")
        
        ws = wb[sheet_name]
        
        print("\n3. Merged cell ranges:")
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        if merged_ranges:
            print(f" Found {len(merged_ranges)} merged ranges (e.g. {', '.join(merged_ranges[:5])}... )")
        else:
            print(" No merged cells found.")
            
        likely_labels = []
        ambiguous_labels = []
        
        print("\n4. Non-empty cells (First 15 printed):")
        printed_count = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    val = str(cell.value).strip()
                    coord = cell.coordinate
                    
                    # Check if cell is in any merged range
                    in_merged = None
                    for merged_range in ws.merged_cells.ranges:
                        if coord in merged_range:
                            in_merged = str(merged_range)
                            break
                            
                    classification = classify_label(val)
                    
                    if printed_count < 15:
                        print(f" - {coord}: '{val}' | Merged: {in_merged} | Class: {classification}")
                        printed_count += 1
                        
                    if classification in ["classroom/lab", "room", "toilet", "stair", "library", "technical_room"]:
                        likely_labels.append((coord, val, classification))
                    else:
                        ambiguous_labels.append((coord, val, classification))
                        
        print("\n5. Likely room/facility labels per floor (First 10):")
        for l in likely_labels[:10]:
            print(f" - {l}")
            
        print("\n6. Ambiguous/problematic labels (First 10):")
        for a in ambiguous_labels[:10]:
            print(f" - {a}")
            
if __name__ == '__main__':
    audit_xlsx()
